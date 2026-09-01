from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Request
from fastapi.responses import FileResponse
from pathlib import Path
from openpyxl import load_workbook
from pypdf import PdfReader
from rapidfuzz import fuzz
import io, csv, re, json

app = FastAPI(title='Stock Statement Compilation Portal')
BASE_DIR = Path(__file__).resolve().parent.parent
@app.get("/", include_in_schema=False)
def home():
    return FileResponse(BASE_DIR / "index.html")

@app.get('/health')
def health():
    return {'ok': True, 'service': 'stock-statement-compiler', 'note': 'Use /api/health and /api/analyze'}

def num(x):
    if x is None: return 0.0
    s=str(x).strip().replace(',','')
    if s in ('','-','—','–'): return 0.0
    try: return float(s)
    except: return 0.0

def sku_norm(s):
    s=str(s or '').upper().replace('’', "'")
    s=s.replace('CILNIREM','CILNIKEM')
    s=re.sub(r'(?<=\d)SMG\b','5MG',s)
    s=s.replace('–','-').replace('—','-')
    s=re.sub(r'\b(?:TABLETS?|TABS?|CAPSULES?|CAPS?|ORAL|ER)\b',' ',s)
    s=re.sub(r"\b\d+(?:X1|X|TAB|TABS|,S|'S|S)\b",' ',s)
    s=re.sub(r'[^A-Z0-9./+]+','',s)
    return s

def norm_identity(s):
    """Normalize HQ/stockist text for reliable comparison."""
    s=str(s or '').upper().replace('’', "'")
    s=s.replace('&',' AND ')
    s=re.sub(r'[^A-Z0-9]+',' ',s)
    s=re.sub(r'\s+',' ',s).strip()
    return s

def norm_customer_identity(s):
    s=norm_identity(s)
    # PDF headers often append the HQ after a comma, e.g. CUSTOMER,AKOLA.
    # Remove common business/location words so master customer names match.
    s=re.sub(r'\b(M S|MS|M S|MESSRS|AKOLA|NAGPUR|AMRAVATI|WASHIM|BULDHANA|YAVATMAL|WARDHA|JALGAON|NASHIK|PUNE|AURANGABAD|CHHATRAPATI SAMBHAJINAGAR|HEADQUARTER|HQ|CORPORATION|CORP|COMPANY|DISTRIBUTOR|DISTRIBUTORS|STOCKIST|AGENCIES|AGENCY|PHARMA|PHARMACEUTICALS)\b',' ',s)
    return re.sub(r'\s+',' ',s).strip()

def extract_pdf_identity(data):
    """Extract stockist + HQ from the first report header in a PDF.

    Known report format:
      KRUSHNA PHARMA,AKOLA
      Stock & Sales Report for the month ...
    """
    reader=PdfReader(io.BytesIO(data))
    header_lines=[]
    for page in reader.pages[:3]:
        text=page.extract_text() or ''
        lines=[' '.join(str(x).replace('\u00a0',' ').split()) for x in text.splitlines()]
        for line in lines:
            if line:
                header_lines.append(line)
        # The identity is normally before the report title on page 1.
        if any(re.search(r'Stock\s*&\s*Sales\s*Report', x, re.I) for x in lines):
            break

    report_idx=next((i for i,x in enumerate(header_lines) if re.search(r'Stock\s*&\s*Sales\s*Report',x,re.I)),None)
    candidates=header_lines[:report_idx] if report_idx is not None else header_lines[:12]

    for line in candidates:
        if re.search(r'Product\s+Name|ALKEM|Stock\s*&\s*Sales',line,re.I):
            continue
        # Expected form: CUSTOMER,HQ. Permit spaces around the comma.
        m=re.match(r'^(.+?)\s*,\s*([A-Za-z][A-Za-z .&()\-]*)$',line)
        if not m:
            continue
        customer=m.group(1).strip(' .')
        hq=m.group(2).strip(' .')
        if len(customer)>=3 and len(hq)>=2:
            return {'customer':customer,'hq':hq,'source':'PDF header','raw_header':line}

    return {'customer':'','hq':'','source':'Not detected','raw_header':''}

def identify_against_master(identity, master_rows):
    """Resolve PDF identity against the uploaded master list."""
    raw_customer=identity.get('customer','')
    raw_hq=identity.get('hq','')
    if not raw_customer or not raw_hq:
        return {**identity,'customer':'','hq':'','confidence':0,'match_status':'NOT DETECTED'}

    ckey=norm_customer_identity(raw_customer)
    hkey=norm_identity(raw_hq)
    best=None
    for row in master_rows:
        mc=str(row.get('CUSTOMER NAME',''))
        mh=str(row.get('HQ',''))
        # Strongly prefer matching both customer and HQ.
        cscore=max(
            fuzz.ratio(ckey,norm_customer_identity(mc)),
            fuzz.WRatio(ckey,norm_customer_identity(mc))
        ) if ckey else 0
        hscore=max(
            fuzz.ratio(hkey,norm_identity(mh)),
            fuzz.WRatio(hkey,norm_identity(mh))
        ) if hkey else 0
        score=round((cscore*0.75)+(hscore*0.25))
        exact=(ckey==norm_customer_identity(mc) and hkey==norm_identity(mh))
        if exact:
            score=100
        if best is None or score>best['score']:
            best={'score':score,'customer':mc,'hq':mh,'cscore':round(cscore),'hscore':round(hscore)}

    if not best or best['score']<70:
        return {**identity,'customer':'','hq':'','confidence':best['score'] if best else 0,'match_status':'NOT MATCHED'}

    status='EXACT' if best['score']>=98 else ('HIGH CONFIDENCE' if best['score']>=90 else 'REVIEW')
    return {
        **identity,
        'customer':best['customer'],
        'hq':best['hq'],
        'confidence':best['score'],
        'customer_confidence':best['cscore'],
        'hq_confidence':best['hscore'],
        'match_status':status
    }

def parse_pdf(data):
    """Parse Alkem Stock & Sales rows without assuming a trailing date/value.

    Standard row layout after Pack is:
      LstSL, Open, Recd, Sales (Secondary), Close, Order, Pend

    Some stockists append LstMove date or Stk.Value after these seven fields.
    Therefore we deliberately take the FIRST seven fields after Pack rather
    than the last numeric fields.
    """
    reader=PdfReader(io.BytesIO(data)); rows=[]
    def is_num_token(t):
        t=str(t or '').strip()
        return t in ('-','—','–') or bool(re.fullmatch(r'-?\d[\d,]*(?:\.\d+)?',t))
    def to_num(t):
        t=str(t or '').strip()
        if t in ('','-','—','–'): return 0.0
        try: return float(t.replace(',',''))
        except: return 0.0
    pack_single=re.compile(r"^(?:\d+(?:\.\d+)?(?:['’]S|['’]TAB|S|TAB|TABS|TAB\.|,TAB|CAP|CAPS|CAP\.)?|\d+X\d+|\d+\*\d+|\*\d+|\d+;S)$",re.I)
    pack_word=re.compile(r"^(?:TAB|TABS|CAP|CAPS|S|TAB\.|CAP\.)$",re.I)
    for page in reader.pages:
        text=page.extract_text() or ''
        for raw in text.splitlines():
            line=' '.join(str(raw).replace('\u00a0',' ').split())
            if not line or re.match(r'^[-_=]{5,}',line): continue
            if re.search(r'Stock\s*&\s*Sales\s+Report|Last Month Sales|Closing Value|Receipt Value|Sales Value|Pend\.Value|^Product\s+Name\s+Pack',line,re.I): continue
            toks=line.split()
            pack_start=pack_end=-1; vals=None
            for k in range(0,len(toks)):
                e=k
                if k+1<len(toks) and pack_word.fullmatch(toks[k+1]) and re.fullmatch(r'\d+(?:\.\d+)?',toks[k]):
                    e=k+1
                elif pack_single.fullmatch(toks[k]):
                    e=k
                else:
                    continue
                after=toks[e+1:]
                if len(after)<7: continue
                seven=after[:7]
                if all(is_num_token(x) for x in seven):
                    pack_start,pack_end=k,e
                    vals=[to_num(x) for x in seven]
                    break
            if pack_start<1 or vals is None or len(vals)!=7: continue
            sku=' '.join(toks[:pack_start]).strip()
            if len(sku)<3: continue
            # [LstSL, Open, Recd, Sales/Secondary, Close, Order, Pend]
            rows.append({'source_sku':sku,'sec':vals[3],'close':vals[4]})
    return rows

def parse_csv(data):
    text=data.decode('utf-8-sig','replace'); out=[]; rows=csv.DictReader(io.StringIO(text))
    for r in rows:
        keys={re.sub(r'[^A-Z]','',str(k).upper()):k for k in r}
        def g(*names):
            for n in names:
                if n in keys: return r.get(keys[n])
            return None
        sku=g('PRODUCTNAME','SKU','PRODUCT','ITEM','PRODUCTDESCRIPTION') or ''
        if sku: out.append({'source_sku':str(sku),'sec':num(g('SALES','SECONDARYUNITS','SECONDARY')),'close':num(g('CLOSE','CLOSING','CLOSINGUNITS'))})
    return out

def parse_xlsx(data):
    wb=load_workbook(io.BytesIO(data),data_only=True,read_only=True); ws=wb[wb.sheetnames[0]]; rows=list(ws.iter_rows(values_only=True))
    if not rows:return []
    headers=[re.sub(r'[^A-Z]','',str(x or '').upper()) for x in rows[0]]
    def idx(*names):
        for n in names:
            if n in headers:return headers.index(n)
        return None
    si=idx('PRODUCTNAME','SKU','PRODUCT','ITEM','PRODUCTDESCRIPTION'); sec_i=idx('SALES','SECONDARYUNITS','SECONDARY'); clo_i=idx('CLOSE','CLOSING','CLOSINGUNITS')
    if si is None: raise ValueError('Could not identify SKU/Product column in Excel')
    out=[]
    for rr in rows[1:]:
        sku=rr[si] if si<len(rr) else ''
        if sku: out.append({'source_sku':str(sku),'sec':num(rr[sec_i]) if sec_i is not None and sec_i<len(rr) else 0,'close':num(rr[clo_i]) if clo_i is not None and clo_i<len(rr) else 0})
    return out

def parse_txt(data):
    text=data.decode('utf-8-sig','replace'); out=[]
    for line in text.splitlines():
        parts=re.split(r'\t|,|\s{2,}',line.strip())
        if len(parts)>=3: out.append({'source_sku':' '.join(parts[:-2]),'sec':num(parts[-2]),'close':num(parts[-1])})
    return out

def parse_statement(data, filename):
    ext=Path(filename).suffix.lower()
    if ext=='.pdf': return parse_pdf(data)
    if ext in ('.xlsx','.xls'): return parse_xlsx(data)
    if ext=='.csv': return parse_csv(data)
    if ext=='.txt': return parse_txt(data)
    raise ValueError(f'Unsupported file type: {ext}')

def match_rows(rows,pool,aliases,threshold):
    out=[]; reviews=[]
    for r in rows:
        key=sku_norm(r['source_sku'])
        chosen=aliases.get(key)
        if chosen:
            m=next((p for p in pool if str(p.get('SKU_NAME',''))==chosen),None)
            if m:
                out.append({**r,'master_sku':chosen,'suggestion':chosen,'pts':num(m.get('PTS')),'confidence':100,'status':'VALIDATED'}); continue
        best=None
        for p in pool:
            name=str(p.get('SKU_NAME','')); score=max(fuzz.ratio(key,sku_norm(name)),fuzz.WRatio(key,sku_norm(name)))
            if best is None or score>best[0]: best=(score,p)
        if best:
            score,p=best; suggestion=str(p.get('SKU_NAME','')); item={**r,'master_sku':suggestion if score>=threshold else '','suggestion':suggestion,'pts':num(p.get('PTS')),'confidence':round(score),'status':'AUTO MATCH' if score>=threshold else 'REVIEW'}
        else: item={**r,'master_sku':'','suggestion':'','pts':0,'confidence':0,'status':'UNMATCHED'}
        if item['master_sku']:
            out.append(item)
        else:
            reviews.append(item)
    return out,reviews

@app.post('/api/identify')
@app.post('/identify')
async def identify_endpoint(request:Request, file:UploadFile=File(None), master_json:str=Form('[]')):
    """Identity fallback. Portal normally sends compact JSON, not the PDF binary."""
    try:
        ct=request.headers.get('content-type','')
        if 'application/json' in ct:
            body=await request.json()
            filename=str(body.get('filename') or '')
            payload=body.get('payload') or {}
            candidates=[]
            txt=payload.get('identity_text')
            if txt: candidates += [x.strip() for x in str(txt).splitlines() if x.strip()][:40]
            for sh in payload.get('identity_sheets') or []:
                for row in (sh.get('rows') or [])[:40]:
                    if isinstance(row,list): candidates += [str(x).strip() for x in row if str(x).strip()]
                    elif str(row).strip(): candidates.append(str(row).strip())
            stem=Path(filename).stem.replace('_',' ').replace('-',' ').strip()
            if stem and re.search(r'[A-Za-z]{3,}',stem) and not re.fullmatch(r'[0-9 ()_-]+',stem): candidates.append(stem)
            raw=' '.join(candidates[:20])
            master_rows=json.loads(body.get('master_json') or '[]')
            if not master_rows:
                master_path=BASE_DIR/'master-template.xlsx'
                if master_path.exists():
                    wb=load_workbook(master_path,data_only=True,read_only=True); ws=wb[wb.sheetnames[0]]; vals=list(ws.iter_rows(values_only=True))
                    if vals:
                        hs=[str(x or '').strip() for x in vals[0]]
                        master_rows=[{hs[i]:r[i] if i<len(r) else '' for i in range(len(hs))} for r in vals[1:]]
            best=None
            for c in candidates:
                ck=norm_customer_identity(c)
                if not ck: continue
                for row in master_rows:
                    mc=str(row.get('CUSTOMER NAME','')); mk=norm_customer_identity(mc)
                    if not mk: continue
                    score=max(fuzz.ratio(ck,mk),fuzz.WRatio(ck,mk))
                    if best is None or score>best['score']: best={'score':round(score),'customer':mc,'hq':str(row.get('HQ','')),'source':c}
            if best and best['score']>=88:
                return {'filename':filename,'customer':best['customer'],'hq':best['hq'],'confidence':best['score'],'match_status':'EXACT' if best['score']>=95 else 'HIGH CONFIDENCE','source':'JSON identity fallback','raw_header':best['source']}
            return {'filename':filename,'customer':'','hq':'','confidence':0,'match_status':'NOT MATCHED','source':'No reliable Master match'}
        if file is None: raise HTTPException(400,'No file supplied')
        data=await file.read()
        if len(data)>4_450_000: raise HTTPException(413,'Statement exceeds Vercel request limit (~4.4 MB).')
        master_rows=json.loads(master_json or '[]')
        identity=extract_pdf_identity(data)
        return identify_against_master(identity,master_rows)
    except HTTPException: raise
    except Exception as e: raise HTTPException(500,f'{type(e).__name__}: {e}')

@app.post('/api/analyze')
@app.post('/analyze')
async def analyze_endpoint(request:Request, file:UploadFile=File(None), pool_json:str=Form('[]'), aliases_json:str=Form('{}'), threshold:int=Form(85)):
    """Accept compact JSON from the browser; retain multipart compatibility."""
    try:
        ct=request.headers.get('content-type','')
        if 'application/json' in ct:
            body=await request.json()
            filename=str(body.get('filename') or 'statement')
            payload=body.get('payload') or {}
            rows=payload.get('rows') or []
            pool=body.get('pool_json') or []
            if isinstance(pool,str): pool=json.loads(pool or '[]')
            aliases=body.get('aliases_json') or {}
            if isinstance(aliases,str): aliases=json.loads(aliases or '{}')
            threshold=int(body.get('threshold') or 85)
        else:
            if file is None: raise HTTPException(400,'No file supplied')
            data=await file.read()
            if len(data)>4_450_000: raise HTTPException(413,'Statement exceeds Vercel request limit (~4.4 MB).')
            filename=file.filename or 'statement'; pool=json.loads(pool_json or '[]'); aliases=json.loads(aliases_json or '{}'); rows=parse_statement(data,filename)
        matched,reviews=match_rows(rows,pool,aliases,threshold)
        return {'file':filename,'rows':len(rows),'sec_total':sum(num(x.get('sec')) for x in rows),'close_total':sum(num(x.get('close')) for x in rows),'matched':matched,'reviews':reviews}
    except HTTPException: raise
    except Exception as e: raise HTTPException(500,f'{type(e).__name__}: {e}')
